"""批量导入解析器单元测试。"""
from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from backend.api.import_export import (
    _coerce_row,
    _parse_characteristics,
    _read_csv,
    _read_xlsx,
)


class TestParseCharacteristics:
    @pytest.mark.parametrize(
        "input_value,expected",
        [
            ("内向", ["内向"]),
            ("内向|喜欢挑战", ["内向", "喜欢挑战"]),
            ("内向、喜欢挑战", ["内向", "喜欢挑战"]),
            ("内向,喜欢挑战", ["内向", "喜欢挑战"]),
            ("内向，喜欢挑战", ["内向", "喜欢挑战"]),
            ("", []),
            (None, []),
        ],
    )
    def test_various_separators(
        self, input_value, expected: list[str]
    ) -> None:
        assert _parse_characteristics(input_value) == expected


class TestCoerceRow:
    def test_basic_mapping(self) -> None:
        row = {
            "姓名": "张三",
            "年龄": "10",
            "性别": "男",
            "基础水平": "初级",
        }
        out = _coerce_row(row)
        assert out["name"] == "张三"
        assert out["age"] == 10
        assert out["gender"] == "男"
        assert out["base_level"] == "初级"

    def test_age_empty_string(self) -> None:
        row = {"姓名": "李四", "年龄": ""}
        out = _coerce_row(row)
        assert out["age"] is None

    def test_age_invalid(self) -> None:
        row = {"姓名": "王五", "年龄": "abc"}
        out = _coerce_row(row)
        assert out["age"] is None

    def test_characteristics_parsed(self) -> None:
        row = {"姓名": "赵六", "性格特点": "内向|喜欢挑战"}
        out = _coerce_row(row)
        assert out["characteristics"] == ["内向", "喜欢挑战"]

    def test_unknown_column_kept(self) -> None:
        row = {"姓名": "测试", "未知列": "value"}
        out = _coerce_row(row)
        assert out["未知列"] == "value"


class TestReadCSV:
    def test_read_simple_csv(self) -> None:
        csv_content = (
            "姓名,年龄,性别,基础水平\n"
            "张三,10,男,入门\n"
            "李四,11,女,初级\n"
        ).encode()
        rows = _read_csv(csv_content)
        assert len(rows) == 2
        assert rows[0]["姓名"] == "张三"
        assert rows[1]["基础水平"] == "初级"

    def test_read_with_bom(self) -> None:
        csv_content = "﻿姓名,年龄\n张三,10\n".encode()
        rows = _read_csv(csv_content)
        assert rows[0]["姓名"] == "张三"

    def test_read_empty(self) -> None:
        rows = _read_csv("姓名,年龄\n".encode())
        assert rows == []


class TestReadXLSX:
    def test_read_simple_xlsx(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "年龄", "基础水平"])
        ws.append(["张三", 10, "入门"])
        ws.append(["李四", 11, "初级"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        rows = _read_xlsx(buf.getvalue())
        assert len(rows) == 2
        assert rows[0]["姓名"] == "张三"
        assert rows[0]["年龄"] == 10
        assert rows[1]["基础水平"] == "初级"

    def test_read_empty_xlsx(self) -> None:
        wb = Workbook()
        ws = wb.active
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        rows = _read_xlsx(buf.getvalue())
        assert rows == []

    def test_skip_blank_rows(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "年龄"])
        ws.append(["张三", 10])
        ws.append([None, None])
        ws.append(["李四", 11])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        rows = _read_xlsx(buf.getvalue())
        # 全空行会被跳过
        assert len(rows) == 2
