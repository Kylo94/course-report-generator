"""
学生批量导入 + Word 导入 API

## 学生批量导入

支持 Excel (.xlsx) 与 CSV 文件格式。

文件格式约定（首行为表头）：
    姓名, 年龄, 性别, 年级, 基础水平, 性格特点, 家长联系方式, 备注
    张三, 10, 男, 三年级, 入门, 内向|喜欢挑战, 13800001111, 转介绍

- 性格特点：使用 | 或 、 分隔多个标签
- 基础水平：入门 / 初级 / 中级（默认入门）
- 必填：姓名

## Word 导入

- POST /api/import/word — 导入 .docx 文件并解析为结构化字段
"""
from __future__ import annotations

import csv
import io
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from backend.db import get_session
from backend.models import Class as ClassModel
from backend.schemas.student import StudentCreate
from backend.services import students as student_svc
from backend.services.docx_importer import DocxImportError, import_docx
from backend.services.students import ClassNotFoundError
from backend.utils.logger import get_logger

log = get_logger(__name__)

router = APIRouter(prefix="/api/import", tags=["import"])


# 字段名映射（中文表头 → 字段名）
COLUMN_MAP: dict[str, str] = {
    "姓名": "name",
    "年龄": "age",
    "性别": "gender",
    "年级": "grade",
    "基础水平": "base_level",
    "性格特点": "characteristics",
    "家长联系方式": "parent_contact",
    "备注": "note",
    "班级ID": "class_id",
    "班级ID(class_id)": "class_id",
    "班级": "class_name",
}


def _parse_characteristics(value: Any) -> list[str]:
    """解析性格特点：支持 | / 、 , 等分隔符。"""
    if value is None or value == "":
        return []
    s = str(value).strip()
    for sep in ("|", "、", ",", ";", "，"):
        if sep in s:
            return [x.strip() for x in s.split(sep) if x.strip()]
    return [s]


def _coerce_row(row: dict[str, Any]) -> dict[str, Any]:
    """将一行原始数据转为 StudentCreate 兼容的字典。"""
    out: dict[str, Any] = {}
    for k, v in row.items():
        key = COLUMN_MAP.get(str(k).strip(), str(k).strip())
        if key == "characteristics":
            out[key] = _parse_characteristics(v)
        elif key in ("age", "class_id"):
            # 数字字段：空字符串视为 None
            if v == "" or v is None:
                out[key] = None
            else:
                try:
                    out[key] = int(v)
                except (ValueError, TypeError):
                    out[key] = None
        else:
            out[key] = str(v).strip() if v is not None else None
    return out


def _read_csv(content: bytes) -> list[dict[str, Any]]:
    """读取 CSV 内容。"""
    text = content.decode("utf-8-sig")  # 自动处理 BOM
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _read_xlsx(content: bytes) -> list[dict[str, Any]]:
    """读取 Excel 内容。"""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    return [
        {h: v for h, v in zip(headers, row) if h}
        for row in rows
        if any(v is not None and str(v).strip() for v in row)
    ]


@router.post(
    "/students",
    response_model=dict,
    summary="批量导入学生（Excel/CSV），支持设置默认值",
)
async def import_students(
    file: UploadFile = File(..., description="Excel(.xlsx) 或 CSV 文件"),
    default_class_id: int | None = Form(None, description="默认班级ID（CSV 中缺失时使用）"),
    default_base_level: str | None = Form(None, description="默认基础水平（CSV 中缺失时使用）"),
    default_grade: str | None = Form(None, description="默认年级（CSV 中缺失时使用）"),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """
    导入学生名单。返回：
    {
        "total": 总行数,
        "success": 成功数,
        "failed": 失败数,
        "errors": [{"row": 行号, "error": "错误信息"}],
        "created_ids": [成功创建的学生 ID 列表]
    }

    可选参数（作为 Form 字段传入）：
    - default_class_id: 当 CSV 行中没有班级ID时，使用此值
    - default_base_level: 当 CSV 行中没有基础水平时，使用此值
    - default_grade: 当 CSV 行中没有年级时，使用此值
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供文件名")

    content = await file.read()
    filename = file.filename.lower()

    # 解析文件
    try:
        if filename.endswith(".xlsx"):
            rows = _read_xlsx(content)
        elif filename.endswith(".csv"):
            rows = _read_csv(content)
        else:
            raise HTTPException(
                status_code=400, detail="仅支持 .xlsx 和 .csv 文件"
            )
    except HTTPException:
        raise
    except Exception as e:
        log.exception("解析文件失败: %s", filename)
        raise HTTPException(status_code=400, detail=f"解析文件失败: {e}")

    log.info("开始导入学生: %s 共 %d 行", filename, len(rows))

    # 预加载班级名称 → ID 映射（含 "班级" 列的 CSV 导出来回）
    class_name_map: dict[str, int] = {}
    class_names = {
        str(row.get("班级", "")).strip()
        for row in rows
        if row.get("班级")
    }
    if class_names:
        stmt = select(ClassModel.id, ClassModel.name).where(ClassModel.name.in_(class_names))
        for row in (await session.execute(stmt)).mappings().all():
            class_name_map[row["name"]] = row["id"]

    success = 0
    failed = 0
    errors: list[dict] = []
    created_ids: list[int] = []

    for idx, raw_row in enumerate(rows, start=2):  # 第 1 行是表头
        try:
            row = _coerce_row(raw_row)
            # 班级名称 → ID（支持 CSV 导出来回）
            if row.get("class_id") is None and row.get("class_name"):
                cid = class_name_map.get(row["class_name"].strip())
                if cid is not None:
                    row["class_id"] = cid
                # 班级名称不在数据库中时，保留 class_id=None
            row.pop("class_name", None)  # StudentCreate 没有此字段
            # 应用默认值（仅当 CSV 行中未设置时）
            if row.get("class_id") is None and default_class_id is not None:
                row["class_id"] = default_class_id
            if not row.get("base_level") and default_base_level is not None:
                row["base_level"] = default_base_level
            if not row.get("grade") and default_grade is not None:
                row["grade"] = default_grade
            # 校验：姓名必填
            if not row.get("name"):
                raise ValueError("姓名为必填项")
            data = StudentCreate(**row)
            student = await student_svc.create_student(session, data)
            created_ids.append(student.id)
            success += 1
        except ValidationError as e:
            failed += 1
            errors.append({"row": idx, "error": _format_validation_error(e)})
        except ClassNotFoundError as e:
            failed += 1
            errors.append({"row": idx, "error": str(e)})
        except ValueError as e:
            failed += 1
            errors.append({"row": idx, "error": str(e)})
        except Exception as e:
            failed += 1
            log.exception("导入第 %d 行失败", idx)
            errors.append({"row": idx, "error": f"未知错误: {e}"})

    log.info(
        "导入完成: total=%d success=%d failed=%d",
        len(rows),
        success,
        failed,
    )
    return {
        "total": len(rows),
        "success": success,
        "failed": failed,
        "errors": errors,
        "created_ids": created_ids,
    }


def _format_validation_error(e: ValidationError) -> str:
    """将 Pydantic 校验错误转为可读字符串。"""
    msgs = []
    for err in e.errors():
        field = ".".join(str(x) for x in err["loc"])
        msgs.append(f"{field}: {err['msg']}")
    return "; ".join(msgs)


@router.post(
    "/word",
    summary="导入 Word 文档并解析为结构化字段",
    response_model=dict,
)
async def import_word(
    file: UploadFile = File(..., description="Word(.docx) 文件"),
) -> dict:
    """导入 .docx 文件并解析为课程报告结构化字段。"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供文件名")

    if not file.filename.lower().endswith(".docx"):
        raise HTTPException(status_code=400, detail="仅支持 .docx 文件")

    try:
        content = await file.read()
        result = import_docx(content)
        return result
    except DocxImportError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        log.exception("Word 导入解析失败: %s", file.filename)
        raise HTTPException(status_code=500, detail=f"解析失败: {e}")
